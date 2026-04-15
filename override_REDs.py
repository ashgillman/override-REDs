#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


STRUCTURE_HEADERS = {
    "structure",
    "structure_name",
    "structure name",
    "name",
    "roi",
    "roi_name",
    "organ",
}

RED_HEADERS = {
    "red",
    "relative_electron_density",
    "relative electron density",
    "electron_density",
    "electron density",
}

IGNORED_MISSING_OVERRIDE_STRUCTURES = {
    "Z1-Bridge",
    "Z10-Couch Support",
    "Z2a-Bridge",
    "Z2b-Bridge",
    "Z3-Bridge",
    "Z4-Couch Support",
    "Z5-Hard-plate",
    "Z6-Couch Support",
    "Z7-Couch Support",
    "Z8-Mattress",
    "target vol. 1",
    "target vol. 2",
    "target vol. 3",
    "target vol. 4",
    "target vol. 5",
    "tumor",
}


@dataclass
class StructureEntry:
    name: str
    name_line: int
    value_line: int
    values: list[str]

    @property
    def is_forced_override(self) -> bool:
        return self.values[0] == "2"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Override Monaco plan RED values. For each structure record, the RED "
            "is treated as the 3rd comma-separated value on the line after the "
            "structure name."
        )
    )
    parser.add_argument("plan", type=Path, help="Input Monaco plan file")
    parser.add_argument(
        "spreadsheet",
        type=Path,
        help="CSV, TSV, or XLSX file containing structure name and RED columns",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("plan.RED-overridden"),
        help="Path for the updated plan file",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional XLSX sheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--dump-current",
        type=Path,
        default=None,
        help="Optional CSV path to export the current structure REDs before updating",
    )
    return parser.parse_args()


def normalize_header(value: str) -> str:
    return " ".join(value.strip().lower().replace("_", " ").split())


def is_int_like(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    if stripped[0] in "+-":
        stripped = stripped[1:]
    return stripped.isdigit()


def extract_structures(lines: list[str]) -> list[StructureEntry]:
    structures: list[StructureEntry] = []
    for index in range(len(lines) - 1):
        name = lines[index].strip()
        value_line = lines[index + 1].strip()
        if not name or "," in name:
            continue

        parts = [part.strip() for part in value_line.split(",")]
        if len(parts) != 5:
            continue
        if not is_int_like(parts[0]) or not is_int_like(parts[1]) or not is_int_like(parts[4]):
            continue

        try:
            float(parts[2])
            float(parts[3])
        except ValueError:
            continue

        structures.append(
            StructureEntry(
                name=name,
                name_line=index,
                value_line=index + 1,
                values=parts,
            )
        )
    return structures


def dump_current_reds(structures: Iterable[StructureEntry], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["structure", "current_red"])
        for entry in structures:
            writer.writerow([entry.name, entry.values[2]])


def load_rows_from_delimited(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        reader = csv.reader(handle, dialect)
        rows = list(reader)

    if not rows:
        raise ValueError(f"Spreadsheet file is empty: {path}")

    headers = [cell.strip() for cell in rows[0]]
    body = rows[1:]
    return [dict(zip(headers, row)) for row in body if any(cell.strip() for cell in row)]


def load_rows_from_xlsx(path: Path, sheet_name: str | None) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading XLSX files requires openpyxl. Convert the spreadsheet to CSV/TSV "
            "or install openpyxl."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Spreadsheet file is empty: {path}")

    headers = ["" if cell is None else str(cell).strip() for cell in rows[0]]
    body = rows[1:]
    result: list[dict[str, str]] = []
    for row in body:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        values = ["" if cell is None else str(cell).strip() for cell in row]
        result.append(dict(zip(headers, values)))
    return result


def choose_columns(rows: list[dict[str, str]]) -> tuple[str, str]:
    if not rows:
        raise ValueError("Spreadsheet has no data rows")

    headers = list(rows[0].keys())
    normalized = {header: normalize_header(header) for header in headers}

    structure_column = next(
        (header for header, norm in normalized.items() if norm in STRUCTURE_HEADERS),
        None,
    )
    red_column = next(
        (header for header, norm in normalized.items() if norm in RED_HEADERS),
        None,
    )

    if structure_column and red_column:
        return structure_column, red_column

    non_empty_headers = [header for header in headers if header.strip()]
    if len(non_empty_headers) >= 2:
        return non_empty_headers[0], non_empty_headers[1]

    raise ValueError(
        "Could not identify structure/RED columns. Use a header row such as "
        "'structure,red'."
    )


def load_red_overrides(path: Path, sheet_name: str | None) -> dict[str, str]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        rows = load_rows_from_delimited(path)
    elif suffix == ".xlsx":
        rows = load_rows_from_xlsx(path, sheet_name)
    else:
        raise ValueError(f"Unsupported spreadsheet format: {path.suffix}")

    structure_column, red_column = choose_columns(rows)
    overrides: dict[str, str] = {}
    for row in rows:
        structure = row.get(structure_column, "").strip()
        red = row.get(red_column, "").strip()
        if not structure or not red:
            continue
        try:
            float(red)
        except ValueError as exc:
            raise ValueError(f"Invalid RED value for '{structure}': {red}") from exc
        overrides[structure] = red

    if not overrides:
        raise ValueError("No usable structure/RED rows were found in the spreadsheet")
    return overrides


def main() -> int:
    args = parse_args()

    lines = args.plan.read_text(encoding="utf-8").splitlines()
    structures = extract_structures(lines)
    if not structures:
        raise ValueError(f"No structure records found in plan file: {args.plan}")

    if args.dump_current:
        dump_current_reds(structures, args.dump_current)

    overrides = load_red_overrides(args.spreadsheet, args.sheet)
    matched = 0

    for entry in structures:
        new_red = overrides.get(entry.name)
        if new_red is None:
            continue
        entry.values[2] = new_red
        lines[entry.value_line] = ",".join(entry.values)
        matched += 1

    args.output.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"Found {len(structures)} structure records in {args.plan}")
    print(f"Applied {matched} RED override(s) from {args.spreadsheet}")
    print(f"Wrote updated plan to {args.output}")

    missing_overrides = sorted(
        entry.name
        for entry in structures
        if entry.is_forced_override
        and entry.name not in overrides
        and entry.name not in IGNORED_MISSING_OVERRIDE_STRUCTURES
    )
    if missing_overrides:
        print("Plan structures without RED overrides:")
        for name in missing_overrides:
            print(f"  {name}")

    # missing = sorted(set(overrides) - {entry.name for entry in structures})
    # if missing:
    #     print("Spreadsheet structures not found in plan:", file=sys.stderr)
    #     for name in missing:
    #         print(f"  {name}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
